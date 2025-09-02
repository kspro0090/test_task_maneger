from flask import render_template, redirect, url_for, request, send_file, current_app
from flask_login import login_required, current_user
from . import bp
from ..models import User, Project, Task, Notification, Tag
from ..utils import get_task_stats, is_ajax_request, ajax_response
from ..extensions import db
from datetime import datetime, timedelta
from sqlalchemy import and_, or_, func, desc
import openpyxl
from openpyxl.styles import Font, PatternFill
import tempfile
import os

@bp.route('/')
@bp.route('/dashboard')
@login_required
def dashboard():
    # Get task statistics
    stats = get_task_stats(user=current_user)
    
    # Get recent tasks for current user
    recent_tasks_query = Task.query
    
    if not current_user.is_admin():
        # For employees, only show tasks from projects they're members of
        from app.models import ProjectMember
        project_ids = db.session.query(ProjectMember.project_id).filter(
            ProjectMember.user_id == current_user.id
        ).subquery()
        recent_tasks_query = recent_tasks_query.filter(Task.project_id.in_(project_ids))
    
    recent_tasks = recent_tasks_query.order_by(desc(Task.updated_at)).limit(10).all()
    
    # Get user's assigned tasks
    my_tasks = Task.query.filter_by(assignee_id=current_user.id).filter(
        Task.status != 'Done'
    ).order_by(Task.due_date.asc().nullslast()).limit(5).all()
    
    # Get overdue tasks
    overdue_tasks_query = Task.query.filter(
        and_(
            Task.due_date < datetime.utcnow(),
            Task.status != 'Done'
        )
    )
    
    if not current_user.is_admin():
        from app.models import ProjectMember
        project_ids = db.session.query(ProjectMember.project_id).filter(
            ProjectMember.user_id == current_user.id
        ).subquery()
        overdue_tasks_query = overdue_tasks_query.filter(Task.project_id.in_(project_ids))
    
    overdue_tasks = overdue_tasks_query.limit(5).all()
    
    # Get recent notifications
    recent_notifications = current_user.notifications.filter_by(is_read=False).order_by(
        desc(Notification.created_at)
    ).limit(5).all()
    
    return render_template('main/dashboard.html',
                         stats=stats,
                         recent_tasks=recent_tasks,
                         my_tasks=my_tasks,
                         overdue_tasks=overdue_tasks,
                         recent_notifications=recent_notifications)

@bp.route('/export/tasks.xlsx')
@login_required
def export_tasks():
    # Get filter parameters from request
    project_id = request.args.get('project_id', type=int)
    status = request.args.get('status')
    priority = request.args.get('priority')
    assignee_id = request.args.get('assignee_id', type=int)
    tag = request.args.get('tag')
    overdue_only = request.args.get('overdue_only', type=bool)
    search = request.args.get('search')
    
    # Build query
    query = Task.query.join(Project)
    
    # Apply filters
    if project_id:
        query = query.filter(Task.project_id == project_id)
    
    if status:
        query = query.filter(Task.status == status)
    
    if priority:
        query = query.filter(Task.priority == priority)
    
    if assignee_id:
        query = query.filter(Task.assignee_id == assignee_id)
    
    if tag:
        query = query.join(Task.tags).filter(Tag.name.contains(tag))
    
    if overdue_only:
        query = query.filter(
            and_(
                Task.due_date < datetime.utcnow(),
                Task.status != 'Done'
            )
        )
    
    if search:
        query = query.filter(
            or_(
                Task.title.contains(search),
                Task.description.contains(search)
            )
        )
    
    # For employees, only show tasks from projects they're members of
    if not current_user.is_admin():
        from app.models import ProjectMember
        project_ids = db.session.query(ProjectMember.project_id).filter(
            ProjectMember.user_id == current_user.id
        ).subquery()
        query = query.filter(Task.project_id.in_(project_ids))
    
    tasks = query.order_by(Task.created_at.desc()).all()
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tasks"
    
    # Define headers
    headers = [
        'شناسه کار', 'پروژه', 'عنوان', 'مسئول', 'وضعیت', 'اولویت',
        'برچسب‌ها', 'تخمین ساعت', 'تاریخ سررسید', 'تاریخ ایجاد',
        'تاریخ آخرین بروزرسانی', 'عقب‌افتاده'
    ]
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Write data
    for row, task in enumerate(tasks, 2):
        ws.cell(row=row, column=1, value=task.id)
        ws.cell(row=row, column=2, value=task.project.name)
        ws.cell(row=row, column=3, value=task.title)
        ws.cell(row=row, column=4, value=task.assignee.full_name if task.assignee else '')
        ws.cell(row=row, column=5, value=task.get_status_display())
        ws.cell(row=row, column=6, value=task.get_priority_display())
        ws.cell(row=row, column=7, value=', '.join([tag.name for tag in task.tags]))
        ws.cell(row=row, column=8, value=task.estimated_hours or '')
        ws.cell(row=row, column=9, value=task.due_date.strftime('%Y-%m-%d %H:%M') if task.due_date else '')
        ws.cell(row=row, column=10, value=task.created_at.strftime('%Y-%m-%d %H:%M'))
        ws.cell(row=row, column=11, value=task.updated_at.strftime('%Y-%m-%d %H:%M'))
        ws.cell(row=row, column=12, value='بله' if task.is_overdue() else 'خیر')
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to temporary file
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb.save(temp_file.name)
    temp_file.close()
    
    # Generate filename with timestamp
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'tasks_export_{timestamp}.xlsx'
    
    def remove_file(response):
        try:
            os.unlink(temp_file.name)
        except Exception:
            pass
        return response
    
    return send_file(
        temp_file.name,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@bp.route('/search')
@login_required
def search():
    query = request.args.get('q', '')
    if not query:
        return render_template('main/search_results.html', results=[], query='')
    
    # Search in tasks
    task_query = Task.query.join(Project).filter(
        or_(
            Task.title.contains(query),
            Task.description.contains(query)
        )
    )
    
    # Search in projects
    project_query = Project.query.filter(
        or_(
            Project.name.contains(query),
            Project.description.contains(query)
        )
    )
    
    # For employees, only show results from projects they're members of
    if not current_user.is_admin():
        from app.models import ProjectMember
        project_ids = db.session.query(ProjectMember.project_id).filter(
            ProjectMember.user_id == current_user.id
        ).subquery()
        
        task_query = task_query.filter(Task.project_id.in_(project_ids))
        project_query = project_query.filter(Project.id.in_(project_ids))
    
    tasks = task_query.limit(20).all()
    projects = project_query.limit(10).all()
    
    results = {
        'tasks': tasks,
        'projects': projects
    }
    
    if is_ajax_request():
        return render_template('main/search_results_partial.html', results=results, query=query)
    
    return render_template('main/search_results.html', results=results, query=query)

@bp.route('/help')
@login_required
def help():
    return render_template('main/help.html')

@bp.route('/about')
@login_required
def about():
    return render_template('main/about.html')